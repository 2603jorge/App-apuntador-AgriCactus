# =============================================================================
#  AgriCactus - App del APUNTADOR  (main.py)
#  v2.2 - Fix recepcion broadcast (multicast lock)
# =============================================================================

import datetime
import json
import os
import socket
import threading

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_DISPONIBLE = True
except Exception:
    EXCEL_DISPONIBLE = False

from kivy.lang import Builder
from kivy.clock import Clock
from kivy.properties import StringProperty, ListProperty
from kivy.uix.screenmanager import Screen, FadeTransition
from kivy.utils import platform
from kivymd.app import MDApp
from kivymd.uix.snackbar import Snackbar
from kivymd.uix.list import TwoLineIconListItem, IconLeftWidget

ARCHIVO_LISTAS     = "apuntador_listas.json"
PUERTO_ANUNCIO_CU  = 45682
PUERTO_CUADRILLERO = 45680
PUERTO_RECEPCION   = 45681
PUERTO_APUNTADOR   = 45683   # Auto-validacion puestos fijos
PUERTO_CONSULTA_EMP     = 45690  # Consulta de empleado por credencial (empleados_server.py)
TIMEOUT_CONSULTA_EMP    = 3.0
REINTENTOS_CONSULTA_EMP = 3


def guardar_listas(listas: list):
    try:
        with open(ARCHIVO_LISTAS, 'w', encoding='utf-8') as f:
            json.dump(listas, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[STORAGE] Error: {e}")

def cargar_listas() -> list:
    if os.path.exists(ARCHIVO_LISTAS):
        try:
            with open(ARCHIVO_LISTAS, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return []

def exportar_csv(listas: list, tipo: str = "avance") -> str:
    try:
        fecha_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre    = f"agricactus_{tipo}_{fecha_str}.csv"

        if platform == 'android':
            try:
                from jnius import autoclass
                PA        = autoclass('org.kivy.android.PythonActivity')
                files_dir = PA.mActivity.getExternalFilesDir(None).getAbsolutePath()
            except Exception:
                files_dir = os.path.expanduser("~")
        else:
            files_dir = os.path.expanduser("~")

        ruta = os.path.join(files_dir, nombre)

        with open(ruta, 'w', encoding='utf-8-sig') as f:
            # ── Seccion 1: PUESTOS FIJOS ──────────────────────────────────────
            f.write("=== PUESTOS FIJOS ===\n")
            f.write(
                "FECHA,CUADRILLA,CUADRILLERO,CREDENCIAL,NOMBRE,"
                "PUESTO_FIJO,HORA_DETECCION,GPS,PRESENTE\n"
            )
            for lista in listas:
                fecha       = lista.get("fecha", "")
                cuadrilla   = lista.get("cuadrilla", "")
                cuadrillero = lista.get("cuadrillero", "").replace('\n', ' ')
                for cred, info in lista.get("trabajadores", {}).items():
                    if info.get("tipo_trabajador") != "FIJO":
                        continue
                    nombre_t  = info.get("nombre", "").replace('\n', ' ')
                    puesto    = info.get("puesto_fijo_desc", "")
                    hora_det  = info.get("hora_deteccion", "")
                    gps       = info.get("gps", "")
                    presente  = "SI" if info.get("validado") else "NO"
                    f.write(
                        f"{fecha},{cuadrilla},{cuadrillero},"
                        f"{cred},{nombre_t},{puesto},"
                        f"{hora_det},{gps},{presente}\n"
                    )

            # ── Seccion 2: JORNALEROS ─────────────────────────────────────────
            f.write("\n=== JORNALEROS ===\n")
            f.write(
                "FECHA,CUADRILLA,CUADRILLERO,TIPO_REPORTE,"
                "CREDENCIAL,NOMBRE,GPS,"
                "HORA_ENTRADA,CUADRO_ENTRADA,ACTIVIDAD_ENTRADA,"
                "HORA_SALIDA_COMIDA,HORA_REGRESO_COMIDA,"
                "CAMBIOS_CUADRO,HORA_SALIDA_FINAL,"
                "TOTAL_PERIODOS,PRESENTE\n"
            )
            for lista in listas:
                fecha       = lista.get("fecha", "")
                cuadrilla   = lista.get("cuadrilla", "")
                cuadrillero = lista.get("cuadrillero", "").replace('\n', ' ')
                tipo_rep    = lista.get("tipo_reporte", "AVANCE")
                for cred, info in lista.get("trabajadores", {}).items():
                    if info.get("tipo_trabajador") == "FIJO":
                        continue
                    nombre_t  = info.get("nombre", "").replace('\n', ' ')
                    gps       = info.get("gps", "")
                    presente  = "SI" if info.get("validado") else "NO"
                    periodos  = info.get("periodos", [])

                    def get_p(tipo_p):
                        for p in periodos:
                            if p.get("tipo") == tipo_p:
                                return p
                        return {}

                    entrada    = get_p("entrada")
                    sal_comida = get_p("salida_comida")
                    reg_comida = get_p("regreso_comida")
                    sal_final  = get_p("salida_final")
                    cambios    = [p for p in periodos if p.get("tipo") == "cambio_cuadro"]
                    cambios_txt = "|".join([
                        f"{c.get('hora','')} {c.get('cuadro','')} {c.get('actividad','')[:10]}"
                        for c in cambios
                    ]) or ""

                    f.write(
                        f"{fecha},{cuadrilla},{cuadrillero},{tipo_rep},"
                        f"{cred},{nombre_t},{gps},"
                        f"{entrada.get('hora','')},"
                        f"{entrada.get('cuadro','')},"
                        f"{entrada.get('actividad','')[:20]},"
                        f"{sal_comida.get('hora','')},"
                        f"{reg_comida.get('hora','')},"
                        f"{cambios_txt},"
                        f"{sal_final.get('hora','')},"
                        f"{len(periodos)},{presente}\n"
                    )

        return ruta
    except Exception as e:
        print(f"[CSV] Error: {e}")
        return ""


def _carpeta_salida() -> str:
    if platform == 'android':
        try:
            from jnius import autoclass
            PA = autoclass('org.kivy.android.PythonActivity')
            return PA.mActivity.getExternalFilesDir(None).getAbsolutePath()
        except Exception:
            return os.path.expanduser("~")
    return os.path.expanduser("~")


def exportar_xlsx(listas: list, tipo: str = "avance") -> str:
    """
    Genera un archivo .xlsx real (no texto), con dos hojas: Puestos Fijos
    y Jornaleros, y lo guarda en el almacenamiento local del dispositivo.
    Requiere la libreria openpyxl (agregala a buildozer.spec: ver nota).
    """
    if not EXCEL_DISPONIBLE:
        print("[EXCEL] openpyxl no esta disponible en este build.")
        return ""
    try:
        fecha_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre    = f"agricactus_{tipo}_{fecha_str}.xlsx"
        ruta      = os.path.join(_carpeta_salida(), nombre)

        wb = Workbook()

        estilo_header_font = Font(bold=True, color="FFFFFF")
        estilo_header_fill = PatternFill(start_color="2D4A1E", end_color="2D4A1E", fill_type="solid")
        estilo_wrap        = Alignment(vertical="center", wrap_text=False)

        def escribir_headers(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = estilo_header_font
                cell.fill = estilo_header_fill
                cell.alignment = estilo_wrap
            ws.freeze_panes = "A2"

        def autoajustar(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                letra = col[0].column_letter
                ws.column_dimensions[letra].width = min(max(max_len + 2, 10), 45)

        # ── Hoja 1: PUESTOS FIJOS ───────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Puestos Fijos"
        escribir_headers(ws1, [
            "Fecha", "Cuadrilla", "Cuadrillero", "Credencial", "Nombre",
            "Puesto Fijo", "Hora Detección", "GPS", "Presente"
        ])
        for lista in listas:
            fecha       = lista.get("fecha", "")
            cuadrilla   = lista.get("cuadrilla", "")
            cuadrillero = lista.get("cuadrillero", "").replace('\n', ' ')
            for cred, info in lista.get("trabajadores", {}).items():
                if info.get("tipo_trabajador") != "FIJO":
                    continue
                ws1.append([
                    fecha, cuadrilla, cuadrillero, cred,
                    info.get("nombre", "").replace('\n', ' '),
                    info.get("puesto_fijo_desc", ""),
                    info.get("hora_deteccion", ""),
                    info.get("gps", ""),
                    "SI" if info.get("validado") else "NO"
                ])
        autoajustar(ws1)

        # ── Hoja 2: JORNALEROS ──────────────────────────────────────────────
        ws2 = wb.create_sheet("Jornaleros")
        escribir_headers(ws2, [
            "Fecha", "Cuadrilla", "Cuadrillero", "Tipo Reporte",
            "Credencial", "Nombre", "GPS",
            "Hora Entrada", "Cuadro Entrada", "Actividad Entrada",
            "Salida Comida", "Regreso Comida",
            "Cambios de Cuadro", "Hora Salida Final",
            "Total Periodos", "Presente"
        ])
        for lista in listas:
            fecha       = lista.get("fecha", "")
            cuadrilla   = lista.get("cuadrilla", "")
            cuadrillero = lista.get("cuadrillero", "").replace('\n', ' ')
            tipo_rep    = lista.get("tipo_reporte", "AVANCE")
            for cred, info in lista.get("trabajadores", {}).items():
                if info.get("tipo_trabajador") == "FIJO":
                    continue
                periodos = info.get("periodos", [])

                def get_p(tipo_p):
                    for p in periodos:
                        if p.get("tipo") == tipo_p:
                            return p
                    return {}

                entrada    = get_p("entrada")
                sal_comida = get_p("salida_comida")
                reg_comida = get_p("regreso_comida")
                sal_final  = get_p("salida_final")
                cambios    = [p for p in periodos if p.get("tipo") == "cambio_cuadro"]
                cambios_txt = " | ".join([
                    f"{c.get('hora','')} {c.get('cuadro','')} {c.get('actividad','')[:15]}"
                    for c in cambios
                ])

                ws2.append([
                    fecha, cuadrilla, cuadrillero, tipo_rep,
                    cred, info.get("nombre", "").replace('\n', ' '), info.get("gps", ""),
                    entrada.get("hora", ""), entrada.get("cuadro", ""), entrada.get("actividad", "")[:30],
                    sal_comida.get("hora", ""), reg_comida.get("hora", ""),
                    cambios_txt, sal_final.get("hora", ""),
                    len(periodos), "SI" if info.get("validado") else "NO"
                ])
        autoajustar(ws2)

        wb.save(ruta)
        return ruta
    except Exception as e:
        print(f"[EXCEL] Error: {e}")
        return ""


KV = '''
#:import FadeTransition kivy.uix.screenmanager.FadeTransition

ScreenManager:
    transition: FadeTransition()
    PantallaInicio:
    PantallaDetalle:
    PantallaProgramacion:
    PantallaAutoservicio:


<PantallaInicio>:
    name: 'inicio'

    MDFloatLayout:
        md_bg_color: 0.94, 0.96, 0.94, 1

        MDFloatLayout:
            size_hint_y: 0.13
            pos_hint: {'x': 0, 'top': 1}
            md_bg_color: 0.18, 0.29, 0.12, 1

            Image:
                source: "logo_agricactus.png"
                size_hint: (0.26, 0.76)
                allow_stretch: True
                keep_ratio: True
                pos_hint: {'center_x': 0.14, 'center_y': 0.5}

            MDLabel:
                text: "APUNTADOR"
                font_style: "H6"
                bold: True
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.96, 0.65, 0.14, 1
                pos_hint: {'center_x': 0.60, 'center_y': 0.62}
                size_hint: (0.68, 0.38)

            MDLabel:
                text: root.fecha_hoy
                font_style: "Caption"
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.78, 0.92, 0.78, 1
                pos_hint: {'center_x': 0.60, 'center_y': 0.26}
                size_hint: (0.68, 0.28)

        MDBoxLayout:
            size_hint_y: 0.004
            pos_hint: {'x': 0, 'top': 0.87}
            md_bg_color: 0.96, 0.65, 0.14, 1

        # Estadisticas
        MDCard:
            size_hint: (0.96, 0.11)
            pos_hint: {'center_x': 0.5, 'top': 0.865}
            elevation: 2
            radius: [10, 10, 10, 10]
            md_bg_color: 1, 1, 1, 1

            MDBoxLayout:
                orientation: 'horizontal'
                padding: '8dp'
                spacing: '2dp'

                MDBoxLayout:
                    orientation: 'vertical'
                    MDLabel:
                        text: root.total_cuadrilleros
                        font_style: "H5"
                        bold: True
                        halign: "center"
                        theme_text_color: "Custom"
                        text_color: 0.18, 0.42, 0.18, 1
                    MDLabel:
                        text: "Cuadrilleros"
                        font_style: "Caption"
                        halign: "center"
                        theme_text_color: "Secondary"

                MDBoxLayout:
                    orientation: 'vertical'
                    MDLabel:
                        text: root.total_jornaleros
                        font_style: "H5"
                        bold: True
                        halign: "center"
                        theme_text_color: "Custom"
                        text_color: 0.96, 0.65, 0.14, 1
                    MDLabel:
                        text: "Jornaleros"
                        font_style: "Caption"
                        halign: "center"
                        theme_text_color: "Secondary"

                MDBoxLayout:
                    orientation: 'vertical'
                    MDLabel:
                        text: root.total_fijos
                        font_style: "H5"
                        bold: True
                        halign: "center"
                        theme_text_color: "Custom"
                        text_color: 0.18, 0.29, 0.55, 1
                    MDLabel:
                        text: "Fijos"
                        font_style: "Caption"
                        halign: "center"
                        theme_text_color: "Secondary"

                MDBoxLayout:
                    orientation: 'vertical'
                    MDLabel:
                        text: root.estado_escucha
                        font_style: "Caption"
                        bold: True
                        halign: "center"
                        theme_text_color: "Custom"
                        text_color: root.color_estado
                    MDLabel:
                        text: "Estado"
                        font_style: "Caption"
                        halign: "center"
                        theme_text_color: "Secondary"

        # Programación del día (importada por QR)
        MDBoxLayout:
            orientation: 'horizontal'
            size_hint: (0.96, 0.065)
            pos_hint: {'center_x': 0.5, 'top': 0.745}
            spacing: '6dp'

            MDRaisedButton:
                text: "📷 IMPORTAR PROGRAMACIÓN"
                md_bg_color: 0.96, 0.65, 0.14, 1
                text_color: 0.12, 0.22, 0.08, 1
                size_hint_x: 0.5
                font_size: '11sp'
                elevation: 2
                on_release: root.importar_programacion()

            MDRaisedButton:
                text: "📋 VER PROGRAMACIÓN"
                md_bg_color: 0.18, 0.29, 0.55, 1
                size_hint_x: 0.5
                font_size: '11sp'
                elevation: 2
                on_release: root.ver_programacion()

        # Autoservicio: checar credencial sin cuadrilla / sin celular del cuadrillero
        MDRectangleFlatIconButton:
            icon: "card-account-details"
            text: "📇 AUTOSERVICIO — CHECAR POR CREDENCIAL"
            theme_text_color: "Custom"
            text_color: 0.18, 0.29, 0.12, 1
            line_color: 0.18, 0.29, 0.12, 1
            pos_hint: {'center_x': 0.5, 'top': 0.678}
            size_hint: (0.96, None)
            height: '38dp'
            font_size: '11sp'
            on_release: app.root.current = 'autoservicio'

        # Lista cuadrilleros
        MDCard:
            size_hint: (0.96, 0.375)
            pos_hint: {'center_x': 0.5, 'top': 0.625}
            elevation: 2
            radius: [10, 10, 10, 10]
            md_bg_color: 1, 1, 1, 1

            MDBoxLayout:
                orientation: 'vertical'
                padding: '4dp'

                MDLabel:
                    text: "Cuadrilleros  [ toca para ver detalle ]"
                    font_style: "Caption"
                    bold: True
                    halign: "center"
                    theme_text_color: "Custom"
                    text_color: 0.18, 0.29, 0.12, 1
                    size_hint_y: None
                    height: '26dp'

                ScrollView:
                    MDList:
                        id: lista_cuadrilleros

        # Botones
        MDBoxLayout:
            orientation: 'horizontal'
            size_hint: (0.96, 0.08)
            pos_hint: {'center_x': 0.5, 'y': 0.12}
            spacing: '6dp'

            MDRaisedButton:
                text: "PEDIR AVANCE"
                md_bg_color: 0.18, 0.29, 0.12, 1
                size_hint_x: 0.5
                elevation: 3
                on_release: root.pedir_todas_listas()

            MDRaisedButton:
                text: "EXPORTAR EXCEL"
                md_bg_color: 0.18, 0.29, 0.55, 1
                size_hint_x: 0.5
                elevation: 3
                on_release: root.exportar()

        MDBoxLayout:
            orientation: 'horizontal'
            size_hint: (0.96, 0.08)
            pos_hint: {'center_x': 0.5, 'y': 0.03}
            spacing: '6dp'

            MDRaisedButton:
                text: "REPORTE FINAL"
                md_bg_color: 0.29, 0.40, 0.25, 1
                size_hint_x: 0.5
                elevation: 3
                on_release: root.exportar_final()

            MDRaisedButton:
                text: "LIMPIAR DIA"
                md_bg_color: 0.96, 0.65, 0.14, 1
                text_color: 0.12, 0.22, 0.08, 1
                size_hint_x: 0.5
                elevation: 3
                on_release: root.limpiar_dia()


<PantallaDetalle>:
    name: 'detalle'

    MDFloatLayout:
        md_bg_color: 0.94, 0.96, 0.94, 1

        MDFloatLayout:
            size_hint_y: 0.13
            pos_hint: {'x': 0, 'top': 1}
            md_bg_color: 0.18, 0.29, 0.12, 1

            MDLabel:
                text: root.titulo_detalle
                font_style: "H6"
                bold: True
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.96, 0.65, 0.14, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.62}
                size_hint: (1, 0.5)

            MDLabel:
                text: root.subtitulo_detalle
                font_style: "Caption"
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.78, 0.92, 0.78, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.26}
                size_hint: (1, 0.3)

        MDBoxLayout:
            size_hint_y: 0.004
            pos_hint: {'x': 0, 'top': 0.87}
            md_bg_color: 0.96, 0.65, 0.14, 1

        MDCard:
            size_hint: (0.96, 0.77)
            pos_hint: {'center_x': 0.5, 'top': 0.865}
            elevation: 2
            radius: [10, 10, 10, 10]
            md_bg_color: 1, 1, 1, 1

            MDBoxLayout:
                orientation: 'vertical'
                padding: '4dp'

                MDLabel:
                    text: root.info_lista
                    font_style: "Caption"
                    halign: "center"
                    theme_text_color: "Custom"
                    text_color: 0.12, 0.22, 0.08, 1
                    size_hint_y: None
                    height: '60dp'

                ScrollView:
                    MDList:
                        id: lista_detalle

        MDRectangleFlatButton:
            text: "REGRESAR"
            theme_text_color: "Custom"
            text_color: 0.18, 0.29, 0.12, 1
            line_color: 0.18, 0.29, 0.12, 1
            size_hint: (0.96, 0.07)
            pos_hint: {'center_x': 0.5, 'y': 0.01}
            on_release: app.root.current = 'inicio'


<PantallaProgramacion>:
    name: 'programacion'

    MDFloatLayout:
        md_bg_color: 0.94, 0.96, 0.94, 1

        MDFloatLayout:
            size_hint_y: 0.13
            pos_hint: {'x': 0, 'top': 1}
            md_bg_color: 0.18, 0.29, 0.12, 1

            MDLabel:
                text: "PROGRAMACIÓN DEL DÍA"
                font_style: "H6"
                bold: True
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.96, 0.65, 0.14, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.62}
                size_hint: (1, 0.5)

            MDLabel:
                text: root.subtitulo_prog
                font_style: "Caption"
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.78, 0.92, 0.78, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.26}
                size_hint: (1, 0.3)

        MDBoxLayout:
            size_hint_y: 0.004
            pos_hint: {'x': 0, 'top': 0.87}
            md_bg_color: 0.96, 0.65, 0.14, 1

        MDCard:
            size_hint: (0.96, 0.77)
            pos_hint: {'center_x': 0.5, 'top': 0.865}
            elevation: 2
            radius: [10, 10, 10, 10]
            md_bg_color: 1, 1, 1, 1

            MDBoxLayout:
                orientation: 'vertical'
                padding: '4dp'

                MDLabel:
                    text: root.sin_datos_texto
                    font_style: "Caption"
                    halign: "center"
                    theme_text_color: "Custom"
                    text_color: 0.6, 0.6, 0.6, 1
                    size_hint_y: None
                    height: '30dp' if root.sin_datos_texto else '0dp'

                ScrollView:
                    MDList:
                        id: lista_programacion

        MDBoxLayout:
            orientation: 'horizontal'
            size_hint: (0.96, 0.07)
            pos_hint: {'center_x': 0.5, 'y': 0.01}
            spacing: '6dp'

            MDRaisedButton:
                text: "🔄 ACTUALIZAR"
                md_bg_color: 0.18, 0.29, 0.12, 1
                size_hint_x: 0.5
                elevation: 3
                on_release: root.actualizar_comparacion()

            MDRectangleFlatButton:
                text: "REGRESAR"
                theme_text_color: "Custom"
                text_color: 0.18, 0.29, 0.12, 1
                line_color: 0.18, 0.29, 0.12, 1
                size_hint_x: 0.5
                on_release: app.root.current = 'inicio'


<PantallaAutoservicio>:
    name: 'autoservicio'

    MDFloatLayout:
        md_bg_color: 0.94, 0.96, 0.94, 1

        MDFloatLayout:
            size_hint_y: 0.13
            pos_hint: {'x': 0, 'top': 1}
            md_bg_color: 0.18, 0.29, 0.12, 1

            MDLabel:
                text: "AUTOSERVICIO"
                font_style: "H6"
                bold: True
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.96, 0.65, 0.14, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.62}
                size_hint: (1, 0.5)

            MDLabel:
                text: "Sin cuadrilla o sin celular del cuadrillero"
                font_style: "Caption"
                halign: "center"
                theme_text_color: "Custom"
                text_color: 0.78, 0.92, 0.78, 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.26}
                size_hint: (1, 0.3)

        MDBoxLayout:
            size_hint_y: 0.004
            pos_hint: {'x': 0, 'top': 0.87}
            md_bg_color: 0.96, 0.65, 0.14, 1

        MDLabel:
            text: "El trabajador escanea o escribe su credencial. Se jala su nombre, actividad y cuadro fijos desde la laptop (empleados_server.py) y queda registrada su asistencia de hoy."
            font_style: "Caption"
            halign: "center"
            theme_text_color: "Custom"
            text_color: 0.4, 0.4, 0.4, 1
            pos_hint: {'center_x': 0.5, 'top': 0.84}
            size_hint: (0.9, 0.12)

        MDRaisedButton:
            text: "📷 ESCANEAR CREDENCIAL"
            md_bg_color: 0.18, 0.42, 0.18, 1
            pos_hint: {'center_x': 0.5, 'top': 0.68}
            size_hint: (0.9, 0.09)
            elevation: 4
            on_release: app.escanear_credencial_autoservicio()

        MDTextField:
            id: input_cred_autoservicio
            hint_text: "O escribe el número de credencial"
            input_filter: "int"
            line_color_focus: 0.18, 0.29, 0.12, 1
            pos_hint: {'center_x': 0.5, 'top': 0.55}
            size_hint_x: 0.9
            on_text_validate: root.registrar_manual()

        MDRaisedButton:
            text: "REGISTRAR ASISTENCIA"
            md_bg_color: 0.96, 0.65, 0.14, 1
            text_color: 0.12, 0.22, 0.08, 1
            pos_hint: {'center_x': 0.5, 'top': 0.46}
            size_hint: (0.9, 0.07)
            on_release: root.registrar_manual()

        MDCard:
            id: card_resultado_auto
            size_hint: (0.92, 0.24)
            pos_hint: {'center_x': 0.5, 'top': 0.37}
            elevation: 2
            radius: [12, 12, 12, 12]
            md_bg_color: 1, 1, 1, 1

            MDLabel:
                id: label_resultado_auto
                text: "Esperando escaneo..."
                font_style: "Body2"
                halign: "center"
                valign: "middle"
                theme_text_color: "Custom"
                text_color: 0.4, 0.4, 0.4, 1
                padding: ('10dp', '10dp')

        MDRectangleFlatButton:
            text: "REGRESAR"
            theme_text_color: "Custom"
            text_color: 0.18, 0.29, 0.12, 1
            line_color: 0.18, 0.29, 0.12, 1
            size_hint: (0.9, 0.07)
            pos_hint: {'center_x': 0.5, 'y': 0.02}
            on_release: app.root.current = 'inicio'
'''


class PantallaInicio(Screen):
    fecha_hoy          = StringProperty("")
    total_cuadrilleros = StringProperty("0")
    total_jornaleros   = StringProperty("0")
    total_fijos        = StringProperty("0")
    estado_escucha     = StringProperty("Iniciando...")
    color_estado       = ListProperty([0.96, 0.65, 0.14, 1])

    def on_enter(self):
        self.fecha_hoy = datetime.datetime.now().strftime("%d/%m/%Y  %H:%M")

    def actualizar_ui(self, cuadrilleros: dict):
        self.ids.lista_cuadrilleros.clear_widgets()
        total_j = 0
        total_f = 0

        for cuadrilla, info in cuadrilleros.items():
            nombre     = info.get('nombre', f"Cuadrilla {cuadrilla}").replace('\n', ' ')
            hora       = info.get('hora_deteccion', '--:--')
            n_trab     = info.get('num_trabajadores', 0)
            n_pres     = info.get('num_presentes', 0)
            n_fijos    = info.get('num_fijos', 0)
            n_jorn     = n_trab - n_fijos
            cerrada    = info.get('cerrada', False)
            tipo_rep   = info.get('tipo_reporte', 'AVANCE')
            total_j   += n_jorn
            total_f   += n_fijos

            icono_n = "lock" if cerrada else "clock-outline"
            icono_c = (0.65, 0.08, 0.08, 1) if cerrada else (0.18, 0.42, 0.18, 1)

            icono = IconLeftWidget(
                icon=icono_n,
                theme_text_color="Custom",
                icon_color=icono_c
            )
            item = TwoLineIconListItem(
                text=f"[b]Cuadrilla {cuadrilla}[/b]  —  {nombre}",
                secondary_text=(
                    f"{hora}  |  {n_pres}/{n_trab} pres  "
                    f"|  {n_fijos} fijos  |  {tipo_rep}"
                ),
                on_release=lambda x, c=cuadrilla: self._ver_detalle(c)
            )
            item.add_widget(icono)
            self.ids.lista_cuadrilleros.add_widget(item)

        self.total_cuadrilleros = str(len(cuadrilleros))
        self.total_jornaleros   = str(total_j)
        self.total_fijos        = str(total_f)

    def _ver_detalle(self, cuadrilla):
        app   = MDApp.get_running_app()
        lista = app.listas_recibidas.get(cuadrilla)
        if not lista:
            Snackbar(text="Presiona PEDIR AVANCE primero").open()
            return
        pd = app.root.get_screen('detalle')
        pd.cargar_detalle(cuadrilla, lista)
        app.root.current = 'detalle'

    def pedir_todas_listas(self):
        app   = MDApp.get_running_app()
        count = 0
        for cuadrilla, info in app.cuadrilleros_detectados.items():
            ip = info.get('ip')
            if ip:
                app.pedir_lista_cuadrillero(ip, cuadrilla)
                count += 1
        if count == 0:
            Snackbar(text="Sin cuadrilleros detectados").open()
        else:
            Snackbar(text=f"Pidiendo avance a {count} cuadrillero(s)...").open()

    def exportar(self):
        app = MDApp.get_running_app()
        if not app.listas_recibidas:
            Snackbar(text="Sin listas para exportar").open()
            return
        listas = list(app.listas_recibidas.values())
        if EXCEL_DISPONIBLE:
            ruta = exportar_xlsx(listas, "avance")
            etiqueta = "Excel"
        else:
            ruta = exportar_csv(listas, "avance")
            etiqueta = "CSV (Excel no disponible en este build)"
        if ruta:
            guardar_listas(listas)
            Snackbar(text=f"{etiqueta}: {os.path.basename(ruta)}").open()
        else:
            Snackbar(text="Error al exportar").open()

    def exportar_final(self):
        app     = MDApp.get_running_app()
        finales = {
            k: v for k, v in app.listas_recibidas.items()
            if v.get('jornada_cerrada', False)
        } or app.listas_recibidas
        if not finales:
            Snackbar(text="Sin listas").open()
            return
        if EXCEL_DISPONIBLE:
            ruta = exportar_xlsx(list(finales.values()), "final")
            etiqueta = "Reporte final (Excel)"
        else:
            ruta = exportar_csv(list(finales.values()), "final")
            etiqueta = "Reporte final (CSV, Excel no disponible)"
        if ruta:
            guardar_listas(list(app.listas_recibidas.values()))
            Snackbar(text=f"{etiqueta}: {os.path.basename(ruta)}").open()
        else:
            Snackbar(text="Error al exportar").open()

    def limpiar_dia(self):
        app = MDApp.get_running_app()
        app.cuadrilleros_detectados = {}
        app.listas_recibidas        = {}
        self.ids.lista_cuadrilleros.clear_widgets()
        self.total_cuadrilleros = "0"
        self.total_jornaleros   = "0"
        self.total_fijos        = "0"
        Snackbar(text="Datos limpiados").open()

    def importar_programacion(self):
        app = MDApp.get_running_app()
        app.escanear_programacion()

    def ver_programacion(self):
        app = MDApp.get_running_app()
        if not app.programacion_actual:
            Snackbar(text="Aún no has importado ninguna programación. Toca IMPORTAR PROGRAMACIÓN.").open()
            return
        pp = app.root.get_screen('programacion')
        pp.actualizar_comparacion()
        app.root.current = 'programacion'


class PantallaDetalle(Screen):
    titulo_detalle    = StringProperty("")
    subtitulo_detalle = StringProperty("")
    info_lista        = StringProperty("")

    def cargar_detalle(self, cuadrilla, lista):
        self.titulo_detalle    = f"Cuadrilla {cuadrilla}"
        nombre_cuad = lista.get('cuadrillero', '').replace('\n', ' ')
        cerrada     = lista.get('jornada_cerrada', False)
        tipo_rep    = lista.get('tipo_reporte', 'AVANCE')
        self.subtitulo_detalle = f"{nombre_cuad}  |  {tipo_rep}"

        trabajadores = lista.get('trabajadores', {})
        presentes    = sum(1 for v in trabajadores.values() if v.get('validado'))
        fijos        = sum(1 for v in trabajadores.values()
                          if v.get('tipo_trabajador') == 'FIJO')
        total        = len(trabajadores)
        estado       = "CERRADA" if cerrada else "EN CURSO"
        hora_rep     = lista.get('hora_reporte', '')

        self.info_lista = (
            f"Cuadro: {lista.get('cuadro','')}  [{estado}]  {hora_rep}\n"
            f"Act: {lista.get('actividad','')[:35]}\n"
            f"Presentes: {presentes}/{total}  Fijos: {fijos}"
        )

        self.ids.lista_detalle.clear_widgets()

        # Separador fijos
        from kivymd.uix.list import OneLineListItem
        self.ids.lista_detalle.add_widget(
            OneLineListItem(
                text="── PUESTOS FIJOS ──",
                theme_text_color="Custom",
            )
        )

        items_ord = sorted(
            trabajadores.items(),
            key=lambda x: (0 if x[1].get('tipo_trabajador') == 'FIJO' else 1)
        )

        separador_puesto = True
        for cred, info in items_ord:
            es_fijo  = info.get('tipo_trabajador', '') == 'FIJO'
            validado = info.get('validado', False)
            nombre   = info.get('nombre', '').replace('\n', ' ')
            periodos = info.get('periodos', [])
            gps      = info.get('gps', '')

            if not es_fijo and separador_puesto:
                self.ids.lista_detalle.add_widget(
                    OneLineListItem(text="── JORNALEROS ──")
                )
                separador_puesto = False

            if es_fijo:
                subtxt = (
                    f"{info.get('puesto_fijo_desc','')}  "
                    f"Det: {info.get('hora_deteccion','')}"
                    + (f"  GPS:{gps}" if gps else "")
                )
                icono_n = "briefcase-check"
                icono_c = (0.18, 0.29, 0.55, 1)
            else:
                per_txt = "  ".join([
                    f"{p.get('tipo','').replace('_',' ')[:4].upper()} {p.get('hora','')}"
                    for p in periodos
                ]) or ('PRESENTE' if validado else 'AUSENTE')
                subtxt  = per_txt + (f"  GPS:{gps}" if gps else "")
                icono_n = "check-circle" if validado else "close-circle"
                icono_c = (0.18, 0.42, 0.18, 1) if validado else (0.72, 0.10, 0.10, 1)

            icono = IconLeftWidget(
                icon=icono_n,
                theme_text_color="Custom",
                icon_color=icono_c
            )
            item = TwoLineIconListItem(
                text=f"Cred. {cred}  —  {nombre}",
                secondary_text=subtxt,
            )
            item.add_widget(icono)
            self.ids.lista_detalle.add_widget(item)


def expandir_cuadros(texto: str) -> set:
    """
    Convierte texto libre de cuadros ('301', '301 y 302', '301 al 304',
    '301, 302, 305') en un set de strings de cuadros individuales, para
    poder comparar contra el campo 'cuadro' que reportan los cuadrilleros.
    """
    if not texto:
        return set()
    t = texto.upper()
    t = t.replace(' AL ', '-').replace(' A ', '-').replace(' Y ', ',')
    resultado = set()
    for parte in t.split(','):
        parte = parte.strip()
        if not parte:
            continue
        if '-' in parte:
            extremos = parte.split('-')
            if len(extremos) == 2:
                try:
                    a, b = int(extremos[0].strip()), int(extremos[1].strip())
                    for n in range(min(a, b), max(a, b) + 1):
                        resultado.add(str(n))
                    continue
                except ValueError:
                    pass
        resultado.add(parte)
    return resultado


class PantallaProgramacion(Screen):
    subtitulo_prog   = StringProperty("")
    sin_datos_texto  = StringProperty("")

    def actualizar_comparacion(self):
        app  = MDApp.get_running_app()
        prog = app.programacion_actual

        if not prog:
            self.subtitulo_prog  = ""
            self.sin_datos_texto = "Sin programación importada todavía."
            self.ids.lista_programacion.clear_widgets()
            return

        self.subtitulo_prog = (
            f"Ing. {prog.get('ingeniero','')}  ·  {prog.get('campo','')}  ·  "
            f"{prog.get('fecha','')}"
        )
        renglones = prog.get('renglones', [])
        if not renglones:
            self.sin_datos_texto = "La programación importada no tiene renglones."
            self.ids.lista_programacion.clear_widgets()
            return

        self.sin_datos_texto = ""
        self.ids.lista_programacion.clear_widgets()

        # Cuenta trabajadores validados por cuadro, juntando todas las
        # cuadrillas recibidas hasta ahora.
        conteo_por_cuadro = {}
        for lista in app.listas_recibidas.values():
            for cred, info in lista.get('trabajadores', {}).items():
                if not info.get('validado'):
                    continue
                periodos = info.get('periodos', [])
                cuadro = periodos[-1].get('cuadro', '') if periodos else ''
                cuadro = str(cuadro).strip().upper()
                if not cuadro or cuadro == 'SIN CUADRO':
                    continue
                conteo_por_cuadro[cuadro] = conteo_por_cuadro.get(cuadro, 0) + 1

        for r in renglones:
            cuadros_set = expandir_cuadros(r.get('cuadros', ''))
            llegaron = sum(conteo_por_cuadro.get(c, 0) for c in cuadros_set)
            programado = int(r.get('cantidad', 0) or 0)
            diferencia = llegaron - programado

            if diferencia == 0:
                icono_n, icono_c = "check-circle", (0.18, 0.42, 0.18, 1)
                estado = "Completo"
            elif diferencia > 0:
                icono_n, icono_c = "arrow-up-bold-circle", (0.96, 0.65, 0.14, 1)
                estado = f"+{diferencia} de más"
            else:
                icono_n, icono_c = "arrow-down-bold-circle", (0.72, 0.10, 0.10, 1)
                estado = f"Faltan {abs(diferencia)}"

            sub = (
                f"{llegaron}/{programado} presentes  —  {estado}"
                + (f"  ·  Sup: {r.get('supervisor')}" if r.get('supervisor') else "")
            )
            icono = IconLeftWidget(
                icon=icono_n,
                theme_text_color="Custom",
                icon_color=icono_c
            )
            item = TwoLineIconListItem(
                text=f"Cuadro {r.get('cuadros','')}  —  {r.get('actividad','')}",
                secondary_text=sub,
            )
            item.add_widget(icono)
            self.ids.lista_programacion.add_widget(item)


class PantallaAutoservicio(Screen):
    def registrar_manual(self):
        cred = self.ids.input_cred_autoservicio.text.strip()
        if not cred:
            Snackbar(text="Escribe la credencial primero").open()
            return
        self.ids.input_cred_autoservicio.text = ""
        app = MDApp.get_running_app()
        app.procesar_autoservicio_credencial(cred)


class ApuntadorAgriCactusApp(MDApp):
    cuadrilleros_detectados = {}
    listas_recibidas        = {}
    programacion_actual     = {}
    _escucha_activa         = False
    _recepcion_activa       = False
    _multicast_lock         = None
    _lock_intentado         = False

    def build(self):
        self.theme_cls.theme_style     = "Light"
        self.theme_cls.primary_palette = "Green"
        controlador = Builder.load_string(KV)
        self._solicitar_permisos()
        self._adquirir_multicast_lock()
        Clock.schedule_once(self._iniciar_servicios, 1.0)
        return controlador

    # ── Solicitud activa de permisos ──────────────────────────────────────────
    # Sin esto, Android nunca muestra el dialogo de permisos y en telefonos
    # recientes (Samsung con "configuracion restringida") es casi imposible
    # otorgarlos a mano desde Ajustes.
    def _solicitar_permisos(self):
        if platform != 'android':
            return
        try:
            from android.permissions import request_permissions, Permission
            permisos = [
                Permission.ACCESS_FINE_LOCATION,
                Permission.ACCESS_COARSE_LOCATION,
                Permission.ACCESS_WIFI_STATE,
                Permission.CHANGE_WIFI_MULTICAST_STATE,
                Permission.CAMERA,
                Permission.WRITE_EXTERNAL_STORAGE,
                Permission.READ_EXTERNAL_STORAGE,
            ]
            try:
                permisos.append(Permission.NEARBY_WIFI_DEVICES)  # Android 13+
            except AttributeError:
                pass
            request_permissions(permisos, self._al_resultado_permisos)
        except Exception as e:
            print(f"[PERMISOS] Error solicitando permisos: {e}")

    def _al_resultado_permisos(self, permisos, resultados):
        print(f"[PERMISOS] Resultado: {dict(zip(permisos, resultados))}")

    # ── Fix de recepcion: el anuncio "CUADRILLERO:..." llega por UDP ─────────
    # broadcast (255.255.255.255). En Android, recibir broadcast por WiFi
    # requiere un multicast lock o el sistema lo descarta en silencio. Sin
    # esto, el Apuntador nunca se entera de la IP del cuadrillero y por lo
    # tanto tampoco puede pedirle el avance ("jalar" la info).
    def _adquirir_multicast_lock(self):
        if self._lock_intentado:
            return
        self._lock_intentado = True
        if platform != 'android':
            return
        try:
            from jnius import autoclass, cast
            PythonActivity = autoclass('org.kivy.android.PythonActivity')
            Context        = autoclass('android.content.Context')
            activity       = PythonActivity.mActivity
            wifi_manager   = activity.getSystemService(Context.WIFI_SERVICE)
            wifi_manager   = cast('android.net.wifi.WifiManager', wifi_manager)
            self._multicast_lock = wifi_manager.createMulticastLock("agricactus_apuntador")
            self._multicast_lock.setReferenceCounted(True)
            self._multicast_lock.acquire()
            print("[WIFI] Multicast lock adquirido correctamente.")
        except Exception as e:
            print(f"[WIFI] No se pudo adquirir multicast lock: {e}")

    def _iniciar_servicios(self, dt):
        self.iniciar_escucha_cuadrilleros()
        self.iniciar_recepcion_listas()
        for lista in cargar_listas():
            cuadrilla = lista.get('cuadrilla', '')
            if cuadrilla:
                self.listas_recibidas[cuadrilla] = lista
        pi = self.root.get_screen('inicio')
        pi.estado_escucha = "Activo"
        pi.color_estado   = [0.18, 0.42, 0.18, 1]
        pi.fecha_hoy      = datetime.datetime.now().strftime("%d/%m/%Y  %H:%M")

    def iniciar_escucha_cuadrilleros(self):
        if self._escucha_activa:
            return
        self._escucha_activa = True
        self._adquirir_multicast_lock()

        def _escuchar():
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                    sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
                    sock.bind(('', PUERTO_ANUNCIO_CU))
                    sock.settimeout(2.0)

                    while self._escucha_activa:
                        try:
                            datos_raw, addr = sock.recvfrom(1024)
                            msg    = datos_raw.decode('utf-8').strip()
                            partes = msg.split(':')

                            if len(partes) >= 3 and partes[0] == 'CUADRILLERO':
                                cuadrilla = partes[1]
                                nombre    = ':'.join(partes[2:])
                                ahora     = datetime.datetime.now().strftime("%H:%M:%S")
                                ip        = addr[0]

                                lista_c   = self.listas_recibidas.get(cuadrilla, {})
                                trab      = lista_c.get('trabajadores', {})
                                n_trab    = len(trab)
                                n_pres    = sum(1 for v in trab.values() if v.get('validado'))
                                n_fijos   = sum(
                                    1 for v in trab.values()
                                    if v.get('tipo_trabajador') == 'FIJO'
                                )

                                ya_existe = cuadrilla in self.cuadrilleros_detectados
                                self.cuadrilleros_detectados[cuadrilla] = {
                                    "nombre":           nombre,
                                    "hora_deteccion":   ahora,
                                    "ip":               ip,
                                    "num_trabajadores": n_trab,
                                    "num_presentes":    n_pres,
                                    "num_fijos":        n_fijos,
                                    "cerrada":          lista_c.get('jornada_cerrada', False),
                                    "tipo_reporte":     lista_c.get('tipo_reporte', 'AVANCE')
                                }
                                if not ya_existe:
                                    Clock.schedule_once(lambda dt: self._actualizar_ui(), 0)

                        except socket.timeout:
                            continue
                        except Exception as e:
                            print(f"[WIFI] Error: {e}")

            except Exception as e:
                print(f"[WIFI] Error cuadrilleros: {e}")
            finally:
                self._escucha_activa = False

        threading.Thread(target=_escuchar, daemon=True).start()

    def iniciar_recepcion_listas(self):
        if self._recepcion_activa:
            return
        self._recepcion_activa = True

        def _escuchar():
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                    sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                    sock.bind(('', PUERTO_RECEPCION))
                    sock.settimeout(2.0)

                    while self._recepcion_activa:
                        try:
                            datos_raw, addr = sock.recvfrom(65535)
                            msg     = datos_raw.decode('utf-8').strip()
                            payload = json.loads(msg)

                            if payload.get('tipo') == 'LISTA_CUADRILLA':
                                cuadrilla    = payload.get('cuadrilla', '')
                                self.listas_recibidas[cuadrilla] = payload

                                trabajadores = payload.get('trabajadores', {})
                                n_trab  = len(trabajadores)
                                n_pres  = sum(1 for v in trabajadores.values() if v.get('validado'))
                                n_fijos = sum(
                                    1 for v in trabajadores.values()
                                    if v.get('tipo_trabajador') == 'FIJO'
                                )

                                if cuadrilla in self.cuadrilleros_detectados:
                                    self.cuadrilleros_detectados[cuadrilla].update({
                                        "num_trabajadores": n_trab,
                                        "num_presentes":    n_pres,
                                        "num_fijos":        n_fijos,
                                        "cerrada":          payload.get('jornada_cerrada', False),
                                        "tipo_reporte":     payload.get('tipo_reporte', 'AVANCE')
                                    })

                                # Auto-validar puestos fijos detectados
                                for cred, info in trabajadores.items():
                                    if (info.get('tipo_trabajador') == 'FIJO' and
                                            info.get('ip')):
                                        self._autovalidar_fijo(cred, info['ip'])

                                Clock.schedule_once(lambda dt: self._actualizar_ui(), 0)
                                tipo = payload.get('tipo_reporte', 'AVANCE')
                                Clock.schedule_once(
                                    lambda dt, c=cuadrilla, p=n_pres, t=n_trab, ti=tipo:
                                    Snackbar(text=f"Lista {ti} C{c}: {p}/{t}").open(), 0
                                )

                        except socket.timeout:
                            continue
                        except json.JSONDecodeError:
                            continue
                        except Exception as e:
                            print(f"[WIFI] Error lista: {e}")

            except Exception as e:
                print(f"[WIFI] Error recepcion: {e}")
            finally:
                self._recepcion_activa = False

        threading.Thread(target=_escuchar, daemon=True).start()

    def _autovalidar_fijo(self, credencial: str, ip: str):
        """Envía SCAN_FIJO al trabajador con puesto fijo para auto-validar."""
        def _enviar():
            try:
                msg = f"SCAN_FIJO:{credencial}"
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                    sock.settimeout(3.0)
                    sock.sendto(msg.encode('utf-8'), (ip, PUERTO_APUNTADOR))
                    try:
                        resp_raw, _ = sock.recvfrom(256)
                        resp = resp_raw.decode('utf-8').strip()
                        if resp.startswith(f"OK_FIJO:{credencial}"):
                            print(f"[APUNTADOR] Fijo {credencial} auto-validado")
                    except socket.timeout:
                        pass
            except Exception as e:
                print(f"[APUNTADOR] Error auto-validar {credencial}: {e}")

        threading.Thread(target=_enviar, daemon=True).start()

    def pedir_lista_cuadrillero(self, ip: str, cuadrilla: str):
        def _pedir():
            try:
                msg = f"PEDIR_LISTA:{cuadrilla}"
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                    sock.settimeout(5.0)
                    sock.sendto(msg.encode('utf-8'), (ip, PUERTO_CUADRILLERO))
            except Exception as e:
                print(f"[WIFI] Error pidiendo lista: {e}")

        threading.Thread(target=_pedir, daemon=True).start()

    def _actualizar_ui(self):
        pi = self.root.get_screen('inicio')
        if self.root.current == 'inicio':
            pi.actualizar_ui(self.cuadrilleros_detectados)

    # ── Escaneo nativo compartido (ZXing) ─────────────────────────────────────
    # Se usa tanto para el QR de programación como para la credencial del
    # autoservicio; self._scan_mode dice cuál de los dos está en curso.
    _scan_mode = None

    def _iniciar_escaneo_zxing(self, prompt, modo):
        if platform != 'android':
            Snackbar(text="El escaneo solo funciona en el dispositivo Android").open()
            return
        try:
            from jnius import autoclass
            from android import activity as android_activity
            IntentIntegrator = autoclass('com.journeyapps.barcodescanner.IntentIntegrator')
            PythonActivity   = autoclass('org.kivy.android.PythonActivity')

            self._scan_mode = modo
            integrador = IntentIntegrator(PythonActivity.mActivity)
            integrador.setDesiredBarcodeFormats([IntentIntegrator.QR_CODE])
            integrador.setPrompt(prompt)
            integrador.setBeepEnabled(True)
            integrador.setBarcodeImageEnabled(False)
            integrador.setOrientationLocked(False)
            android_activity.bind(on_activity_result=self._resultado_scan)
            integrador.initiateScan()
        except Exception as e:
            print(f"[SCAN] Error iniciando escaneo: {e}")
            Snackbar(text=f"No se pudo abrir el escaner: {e}").open()

    # ── Importar programación del día por QR (escaneo nativo, sin WebView) ───
    def escanear_programacion(self):
        self._iniciar_escaneo_zxing("Escanea el QR de programación del ingeniero", "programacion")

    def escanear_credencial_autoservicio(self):
        self._iniciar_escaneo_zxing("Escanea la credencial del trabajador", "credencial")

    def _resultado_scan(self, requestCode, resultCode, intent):
        try:
            from jnius import autoclass
            IntentIntegrator = autoclass('com.journeyapps.barcodescanner.IntentIntegrator')
            resultado = IntentIntegrator.parseActivityResult(requestCode, resultCode, intent)
            if resultado is None:
                return
            contenido = resultado.getContents()
            if contenido is None:
                Snackbar(text="Escaneo cancelado").open()
                return
            if self._scan_mode == "credencial":
                credencial = ''.join(ch for ch in contenido if ch.isdigit())
                self.procesar_autoservicio_credencial(credencial)
            else:
                self._procesar_programacion_escaneada(contenido)
        except Exception as e:
            print(f"[SCAN] Error leyendo resultado: {e}")
            Snackbar(text=f"Error leyendo el escaneo: {e}").open()

    def _procesar_programacion_escaneada(self, texto: str):
        try:
            datos = json.loads(texto)
        except Exception:
            Snackbar(text="Ese QR no contiene una programación válida").open()
            return

        if 'renglones' not in datos:
            Snackbar(text="Ese QR no es de programación (formato no reconocido)").open()
            return

        self.programacion_actual = datos
        n = len(datos.get('renglones', []))
        Snackbar(
            text=f"Programación importada: {n} renglón(es) — {datos.get('ingeniero','')}"
        ).open()
        pp = self.root.get_screen('programacion')
        pp.actualizar_comparacion()
        self.root.current = 'programacion'

    # ── AUTOSERVICIO: registro por credencial, sin cuadrillero de por medio ───
    # Consulta empleados_server.py (misma laptop/protocolo que usa el
    # Trabajador) y guarda el registro dentro de una "cuadrilla" sintetica
    # llamada AUTOSERVICIO, para que el CSV y el detalle la incluyan gratis.
    def buscar_empleado_red(self, credencial, callback_ok, callback_error):
        credencial = str(credencial).strip()

        def _worker():
            try:
                with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as sock:
                    sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
                    sock.settimeout(TIMEOUT_CONSULTA_EMP)
                    mensaje = f"CONSULTA_EMP|{credencial}".encode('utf-8')

                    for _ in range(REINTENTOS_CONSULTA_EMP):
                        try:
                            sock.sendto(mensaje, ('255.255.255.255', PUERTO_CONSULTA_EMP))
                            datos_raw, addr = sock.recvfrom(4096)
                            msg    = datos_raw.decode('utf-8').strip()
                            partes = msg.split('|')

                            if partes[0] == 'EMP_OK' and len(partes) >= 2 and partes[1] == credencial:
                                resultado = {
                                    "nombre":       partes[2] if len(partes) > 2 else "",
                                    "nss":          partes[3] if len(partes) > 3 else "",
                                    "cuadrilla":    partes[4] if len(partes) > 4 else "",
                                    "puesto_clave": partes[5] if len(partes) > 5 else "",
                                    "puesto_desc":  partes[6] if len(partes) > 6 else "",
                                }
                                Clock.schedule_once(lambda dt: callback_ok(resultado), 0)
                                return

                            if partes[0] == 'EMP_NOTFOUND' and len(partes) >= 2 and partes[1] == credencial:
                                Clock.schedule_once(
                                    lambda dt: callback_error(
                                        f"No existe ningun empleado con credencial {credencial}"
                                    ), 0
                                )
                                return
                        except socket.timeout:
                            continue

                    Clock.schedule_once(
                        lambda dt: callback_error(
                            "Sin respuesta del servidor. Verifica que la laptop "
                            "este encendida y conectada al mismo WiFi."
                        ), 0
                    )
            except Exception as e:
                Clock.schedule_once(lambda dt, err=str(e): callback_error(f"Error de red: {err}"), 0)

        threading.Thread(target=_worker, daemon=True).start()

    def procesar_autoservicio_credencial(self, credencial):
        credencial = str(credencial).strip()
        if not credencial:
            return

        pa = self.root.get_screen('autoservicio')
        pa.ids.label_resultado_auto.text = f"Buscando credencial {credencial}..."
        pa.ids.label_resultado_auto.text_color = (0.5, 0.5, 0.5, 1)

        lista_auto = self.listas_recibidas.setdefault("AUTOSERVICIO", {
            "tipo": "LISTA_CUADRILLA",
            "fecha": datetime.datetime.now().strftime("%Y-%m-%d"),
            "hora_reporte": datetime.datetime.now().strftime("%H:%M:%S"),
            "cuadrilla": "AUTOSERVICIO",
            "cuadrillero": "Registro individual",
            "cuadro": "",
            "actividad": "",
            "jornada_cerrada": False,
            "tipo_reporte": "AVANCE",
            "trabajadores": {}
        })

        if credencial in lista_auto["trabajadores"]:
            info_previa = lista_auto["trabajadores"][credencial]
            pa.ids.label_resultado_auto.text = (
                f"⚠ {info_previa.get('nombre','')} ya se registró hoy "
                f"a las {info_previa.get('hora_deteccion','')}"
            )
            pa.ids.label_resultado_auto.text_color = (0.7, 0.2, 0.15, 1)
            return

        def _ok(datos: dict):
            ahora = datetime.datetime.now().strftime("%H:%M:%S")
            nombre       = datos.get("nombre", "") or f"Credencial {credencial}"
            puesto_clave = datos.get("puesto_clave", "")
            puesto_desc  = datos.get("puesto_desc", "")
            cuadrilla    = datos.get("cuadrilla", "")

            lista_auto["trabajadores"][credencial] = {
                "nombre":            nombre,
                "hora_deteccion":    ahora,
                "validado":          True,
                "ip":                "",
                "gps":               "",
                "confirmaciones":    0,
                "tipo_trabajador":   "FIJO" if puesto_desc else "JORNALERO",
                "puesto_fijo_clave": puesto_clave,
                "puesto_fijo_desc":  puesto_desc,
                "periodos": [{
                    "tipo": "entrada", "hora": ahora,
                    "cuadro": cuadrilla, "actividad": puesto_desc
                }]
            }
            lista_auto["hora_reporte"] = ahora

            self.cuadrilleros_detectados["AUTOSERVICIO"] = {
                "nombre":           "Registro individual",
                "hora_deteccion":   ahora,
                "ip":               "",
                "num_trabajadores": len(lista_auto["trabajadores"]),
                "num_presentes":    sum(1 for v in lista_auto["trabajadores"].values() if v.get("validado")),
                "num_fijos":        sum(1 for v in lista_auto["trabajadores"].values() if v.get("tipo_trabajador") == "FIJO"),
                "cerrada":          False,
                "tipo_reporte":     "AVANCE"
            }
            self._actualizar_ui()

            pa.ids.label_resultado_auto.text = (
                f"✓ {nombre}\n"
                f"{puesto_desc if puesto_desc else 'Sin actividad fija'}"
                f"{('  ·  Cuadrilla ' + cuadrilla) if cuadrilla else ''}\n"
                f"Registrado a las {ahora}"
            )
            pa.ids.label_resultado_auto.text_color = (0.18, 0.42, 0.18, 1)
            Snackbar(text=f"Asistencia registrada: {nombre}").open()

        def _error(mensaje: str):
            pa.ids.label_resultado_auto.text = mensaje
            pa.ids.label_resultado_auto.text_color = (0.7, 0.2, 0.15, 1)
            Snackbar(text=mensaje).open()

        self.buscar_empleado_red(credencial, _ok, _error)

    def on_stop(self):
        self._escucha_activa   = False
        self._recepcion_activa = False
        if self._multicast_lock is not None:
            try:
                self._multicast_lock.release()
            except Exception:
                pass


if __name__ == '__main__':
    ApuntadorAgriCactusApp().run()

